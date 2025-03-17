#! python3
# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------
# cashbook.py
# cashbookのエントリポイント
# :e ++enc=utf8
# -----------------------------------------------------------------------------
import os
import sys
from dataclasses import dataclass

import openpyxl


# ----------------------------------
# CashBook
# ----------------------------------
@dataclass
class CashBook:
    file_path: str
    target_sheet_name: str
    kamoku_sheet_name: str
    start_row: int
    end_row: int


# ----------------------------------
# disp_useage
# ----------------------------------
def disp_useage() -> str:
    msg: str = '''現金出納帳を集計するツールです

Usage: cashbook.exe ファイル名 シート名 開始行 終了行

OPTION:
    現金出納帳のファイル名
    現金出納帳のシート名（例: R06）
    集計開始行（例: 5）
    集計終了行（例: 160）
'''
    print(msg, file=sys.stderr)
    return msg


# ----------------------------------
# sum_income
# ----------------------------------
def sum_income(cb: CashBook, kamoku: str) -> int:
    kamoku_total: str = 'ALL'
    workbook: openpyxl.Workbook = openpyxl.load_workbook(cb.file_path)
    sheet: openpyxl.worksheet.worksheet.Worksheet = workbook[cb.target_sheet_name]

    total: int = 0
    for row in range(cb.start_row, cb.end_row):
        cell_value: int = sheet[f'H{row}'].value
        if isinstance(cell_value, (int, float)):
            if kamoku == kamoku_total:
                total += cell_value
            elif kamoku == sheet[f'E{row}'].value:
                total += cell_value
            else:
                pass

    return total


# ----------------------------------
# sum_expenses
# ----------------------------------
def sum_expenses(cb: CashBook, kamoku: str) -> int:
    kamoku_total: str = 'ALL'
    workbook: openpyxl.Workbook = openpyxl.load_workbook(cb.file_path)
    sheet: openpyxl.worksheet.worksheet.Worksheet = workbook[cb.target_sheet_name]

    total: int = 0
    for row in range(cb.start_row, cb.end_row):
        cell_value: int = sheet[f'I{row}'].value
        if isinstance(cell_value, (int, float)):
            if kamoku == kamoku_total:
                total += cell_value
            elif kamoku == sheet[f'E{row}'].value:
                total += cell_value
            else:
                pass

    return total


# ----------------------------------
# total_income
# ----------------------------------
def total_income(cb: CashBook) -> int:
    kamoku: str = 'ALL'
    total: int = sum_income(cb, kamoku)
    return total


# ----------------------------------
# total_expenses
# ----------------------------------
def total_expenses(cb: CashBook) -> int:
    kamoku: str = 'ALL'
    total: int = sum_expenses(cb, kamoku)
    return total


# ----------------------------------
# kaihi_income
# ----------------------------------
def kaihi_income(cb: CashBook) -> int:
    kamoku: str = '会費'
    total: int = sum_income(cb, kamoku)
    return total


# ----------------------------------
# kaihi_expenses
# ----------------------------------
def kaihi_expenses(cb: CashBook) -> int:
    kamoku: str = '会費'
    total: int = sum_expenses(cb, kamoku)
    return total


# ----------------------------------
# income_kamoku_list
# ----------------------------------
def income_kamoku_list(cb: CashBook) -> list:
    workbook: openpyxl.Workbook = openpyxl.load_workbook(cb.file_path)
    sheet: openpyxl.worksheet.worksheet.Worksheet = workbook[cb.kamoku_sheet_name]
    k_list: list[str] = []

    delimit_count: int = 0
    is_income: bool = False

    for row in range(2, 50):
        kamoku: str = sheet[f'A{row}'].value
        if kamoku is not None:
            if kamoku.find('<<') == -1:
                if is_income:
                    k_list.append(kamoku)
            else:
                delimit_count += 1
                is_income = True
                if delimit_count == 2:
                    break

    return k_list


# ----------------------------------
# expenses_kamoku_list
# ----------------------------------
def expenses_kamoku_list(cb: CashBook) -> list:
    workbook: openpyxl.Workbook = openpyxl.load_workbook(cb.file_path)
    sheet: openpyxl.worksheet.worksheet.Worksheet = workbook[cb.kamoku_sheet_name]
    k_list: list[str] = []

    delimit_count: int = 0
    is_income: bool = False

    for row in range(2, 50):
        kamoku: str = sheet[f'A{row}'].value
        if kamoku is not None:
            if kamoku.find('<<') == -1:
                if is_income:
                    k_list.append(kamoku)
            else:
                delimit_count += 1
                if delimit_count == 2:
                    is_income = True

    return k_list


# ----------------------------------
# kamoku_income
# ----------------------------------
def kamoku_income(cb: CashBook) -> int:
    # 前年度繰越金は計算で求める
    k_list: list[str] = income_kamoku_list(cb)
    total: int = 0
    sub_total: int = 0

    for kamoku in k_list:
        sub_total = sum_income(cb, kamoku)
        total += sub_total
        print(f'{kamoku}: {sub_total}')

    zennendo: int = total_income(cb) - total
    print(f'前年度繰越金: {zennendo}')

    return total + zennendo


# ----------------------------------
# kamoku_expenses
# ----------------------------------
def kamoku_expenses(cb: CashBook) -> int:
    k_list: list[str] = expenses_kamoku_list(cb)
    total: int = 0
    sub_total: int = 0

    for kamoku in k_list:
        sub_total = sum_expenses(cb, kamoku)
        total += sub_total
        print(f'{kamoku}: {sub_total}')

    return total


# ----------------------------------
# remove_data_validation
# ----------------------------------
def remove_data_validation(cb: CashBook) -> None:
    # データの入力規則を削除（openpyxlは未対応）
    workbook: openpyxl.Workbook = openpyxl.load_workbook(cb.file_path)
    sheet: openpyxl.worksheet.worksheet.Worksheet = workbook[cb.kamoku_sheet_name]

    for data_validation in sheet.data_validations.dataValidation:
        sheet.data_validations.remove(data_validation)

    workbook.save(cb.file_path)


# ----------------------------------
# collect_main
# ----------------------------------
def collect_main(file_path: str, target_sheet_name: str, start_row: int, end_row: int) -> int:
    cb: CashBook = CashBook(file_path, target_sheet_name, '科目', start_row, end_row)

    # データの入力規則を削除
    # remove_data_validation(cb)

    # 集計処理
    total_i: int = total_income(cb)
    total_e: int = total_expenses(cb)

    print('総額の部')
    print(f'収入総額: {total_i}')
    print(f'支出総額: {total_e}')
    print(f'次年度繰越金: {total_i - total_e}')
    print('')

    # kaihi_i: int = kaihi_income(cb)
    # kaihi_e: int = kaihi_expenses(cb)
    # print(f'会費合計: {kaihi_i}')
    # print(f'会費支出: {kaihi_e}')
    # print(f'会費差引: {kaihi_i - kaihi_e}')

    print('収入の部')
    # print(f'収入科目リスト: {income_kamoku_list(cb)}')
    print(f'小計: {kamoku_income(cb)}')
    print('')

    print('支出の部')
    # print(f'支出科目リスト: {expenses_kamoku_list(cb)}')
    print(f'小計: {kamoku_expenses(cb)}')
    print('')

    return 0


# ----------------------------------
# main
# ----------------------------------
if __name__ == '__main__':
    args = sys.argv
    if len(args) == 1:
        disp_useage()
        sys.exit(1)

    if len(args) > 6:
        disp_useage()
        sys.exit(1)

    file_path: str = args[1]
    target_sheet_name: str = args[2]

    # startとendの範囲を指定する
    # for row in range(5, 160):  # H5からH159まで
    start_row: int = int(args[3])
    end_row: int = int(args[4])

    rc: int = collect_main(file_path, target_sheet_name, start_row, end_row)
    os._exit(rc)


# vim:fenc=utf-8 bomb tw=0:
